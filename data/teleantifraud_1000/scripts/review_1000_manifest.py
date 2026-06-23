from __future__ import annotations

import csv
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
INPUT_CSV = ROOT / "teleantifraud_pilot" / "outputs" / "sample_1000" / "sample_manifest_1000_paraformer_hotword_transcribed.csv"
OUTPUT_DIR = ROOT / "teleantifraud_pilot" / "outputs" / "sample_1000"
OUTPUT_CSV = OUTPUT_DIR / "teleantifraud_1000_ai_review.csv"
OUTPUT_XLSX = OUTPUT_DIR / "teleantifraud_1000_ai_review.xlsx"
SUMMARY_JSON = OUTPUT_DIR / "teleantifraud_1000_ai_review_summary.json"


SCENE_MAP = {
    "冒充公检法": "SC-01 冒充公检法",
    "投资理财诈骗": "SC-02 投资理财诈骗",
    "网络刷单诈骗": "SC-03 网络刷单诈骗",
    "保健品骗局": "SC-04 保健品骗局",
    "虚假中奖": "SC-05 虚假中奖",
    "冒充子女": "SC-06 冒充子女",
    "物流快递诈骗": "SC-07 物流快递诈骗",
    "冒充客服退款": "SC-08 冒充客服退款",
    "虚假贷款诈骗": "SC-09 虚假贷款诈骗",
    "虚假征信修复": "SC-10 虚假征信修复",
    "医保社保诈骗": "SC-11 医保社保诈骗",
    "冒充熟人借钱": "SC-12 冒充熟人借钱",
    "银行账户异常诈骗": "SC-13 银行账户异常诈骗",
    "绑架勒索诈骗": "SC-14 绑架勒索诈骗",
}


def read_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_rows(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(rows[0].keys()) if rows else []
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def clean_text(text: str) -> str:
    text = (text or "").strip()
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"https?://\S+", "【模拟链接】", text)
    text = re.sub(r"\b1[3-9]\d{9}\b", "【模拟手机号】", text)
    return text


def has_any(text: str, words: list[str]) -> bool:
    return any(word in text for word in words)


def suggest_label(raw_label: str, text: str) -> tuple[str, str]:
    text = text or ""
    raw_label = raw_label or ""

    if has_any(text, ["公安", "警官", "检察", "法院", "洗钱", "涉案", "安全账户", "办案"]):
        return "公检法/身份核验诈骗", SCENE_MAP["冒充公检法"]
    if has_any(text, ["绑架", "赎金", "人质", "不要报警", "安全就按我说", "你家人现在"]):
        return "绑架勒索诈骗", SCENE_MAP["绑架勒索诈骗"]
    if has_any(text, ["包裹", "快递", "海关", "物流", "违禁", "运单", "理赔"]) and not has_any(text, ["商品", "订单"]):
        return "物流/快递诈骗", SCENE_MAP["物流快递诈骗"]
    if has_any(text, ["退款", "理赔", "订单", "售后", "会员扣费", "商家客服", "平台客服"]):
        return "电商客服退款诈骗", SCENE_MAP["冒充客服退款"]
    if has_any(text, ["贷款", "信贷", "放款", "无抵押", "低息", "额度", "借款", "备用金"]):
        if has_any(text, ["征信", "关闭账户", "注销", "校园贷", "影响信用"]):
            return "虚假征信/贷款账户诈骗", SCENE_MAP["虚假征信修复"]
        if has_any(text, ["APP", "链接", "二维码", "下载"]):
            return "贷款/APP或钓鱼链接诈骗", SCENE_MAP["虚假贷款诈骗"]
        return "虚假贷款诈骗", SCENE_MAP["虚假贷款诈骗"]
    if has_any(text, ["理财", "投资", "收益", "保本", "股票", "基金", "老师带", "年化", "内部"]):
        return "投资理财诈骗", SCENE_MAP["投资理财诈骗"]
    if has_any(text, ["中奖", "彩票", "奖金", "特等奖", "领奖", "保证金", "手续费", "公证费"]):
        return "彩票中奖诈骗", SCENE_MAP["虚假中奖"]
    if has_any(text, ["医保", "社保", "补贴", "社保卡", "医保卡"]):
        return "医保社保诈骗", SCENE_MAP["医保社保诈骗"]
    if has_any(text, ["爸", "妈", "叔叔", "阿姨", "朋友", "亲戚", "临时号码", "手机坏了"]):
        return "冒充熟人/亲友紧急诈骗", SCENE_MAP["冒充熟人借钱"]
    if has_any(text, ["账户异常", "异常交易", "银行卡", "身份信息", "身份证", "验证码", "信用卡", "银行客服"]):
        return "银行账户异常/身份验证诈骗", SCENE_MAP["银行账户异常诈骗"]

    if "投资诈骗" in raw_label:
        return "投资理财诈骗", SCENE_MAP["投资理财诈骗"]
    if "彩票诈骗" in raw_label:
        return "彩票中奖诈骗", SCENE_MAP["虚假中奖"]
    if "银行诈骗" in raw_label or "身份盗窃" in raw_label or "钓鱼诈骗" in raw_label:
        return "银行账户异常/身份验证诈骗", SCENE_MAP["银行账户异常诈骗"]
    if "绑架诈骗" in raw_label:
        return "绑架勒索诈骗", SCENE_MAP["绑架勒索诈骗"]
    if "客服诈骗" in raw_label:
        return "客服/贷款或退款诈骗", "待人工细分"
    return "待人工细分", "待人工细分"


def quality_review(text: str, status: str, suggested_label: str) -> tuple[str, int, str]:
    text_len = len(text or "")
    notes: list[str] = []
    if status != "done":
        return "不可用", 1, f"ASR状态为{status}"
    if text_len < 80:
        notes.append("文本较短")
    if text_len > 900:
        notes.append("文本较长，建议摘要后接入")
    if "待人工细分" in suggested_label:
        notes.append("类型不够明确")
    if has_any(text, ["听不清", "无法识别", "音乐", "噪声"]):
        notes.append("可能存在识别噪声")
    if has_any(text, ["验证码", "银行卡", "身份证", "链接", "APP", "转账", "保证金", "手续费", "账户", "贷款", "收益"]):
        notes.append("包含典型风险触发词")

    if text_len >= 160 and "类型不够明确" not in notes:
        score = 4
        conclusion = "可直接用"
    elif text_len >= 80:
        score = 3
        conclusion = "需轻微清洗"
    else:
        score = 2
        conclusion = "仅作台账"

    if text_len >= 260 and "包含典型风险触发词" in notes and conclusion == "可直接用":
        score = 5
    if "文本较长，建议摘要后接入" in notes:
        conclusion = "需轻微清洗"
        score = min(score, 4)

    return conclusion, score, "；".join(notes) if notes else "语义基本完整"


def build_review() -> list[dict[str, Any]]:
    reviewed: list[dict[str, Any]] = []
    for row in read_rows(INPUT_CSV):
        text = clean_text(row.get("paraformer_hotword_text", ""))
        raw_label = row.get("fraud_types", "")
        suggested_label, mapped_scene = suggest_label(raw_label, text)
        conclusion, score, notes = quality_review(text, row.get("paraformer_hotword_status", ""), suggested_label)
        reviewed.append({
            "样本ID": row.get("sample_id", ""),
            "原始标签": raw_label,
            "建议细分标签": suggested_label,
            "映射训练场景": mapped_scene,
            "音频路径": row.get("audio_path", ""),
            "ASR状态": row.get("paraformer_hotword_status", ""),
            "文本长度": len(text),
            "复核结论": conclusion,
            "质量评分(1-5)": score,
            "是否建议接入系统": "是" if conclusion in {"可直接用", "需轻微清洗"} and mapped_scene != "待人工细分" else "否",
            "转写全文": text,
            "备注/问题": notes,
        })
    return reviewed


def write_xlsx(rows: list[dict[str, Any]]) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "1000条AI初审台账"
    headers = list(rows[0].keys()) if rows else []
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "A": 12, "B": 14, "C": 24, "D": 24, "E": 42, "F": 10,
        "G": 10, "H": 14, "I": 12, "J": 16, "K": 90, "L": 38,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    for row_cells in sheet.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    summary = workbook.create_sheet("统计摘要")
    counters = {
        "复核结论": Counter(row["复核结论"] for row in rows),
        "建议细分标签": Counter(row["建议细分标签"] for row in rows),
        "映射训练场景": Counter(row["映射训练场景"] for row in rows),
        "是否建议接入系统": Counter(row["是否建议接入系统"] for row in rows),
    }
    current_row = 1
    for title, counter in counters.items():
        summary.cell(current_row, 1, title).font = Font(bold=True)
        current_row += 1
        summary.cell(current_row, 1, "类别").font = Font(bold=True)
        summary.cell(current_row, 2, "数量").font = Font(bold=True)
        current_row += 1
        for key, value in counter.most_common():
            summary.cell(current_row, 1, key)
            summary.cell(current_row, 2, value)
            current_row += 1
        current_row += 2
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 12

    workbook.save(OUTPUT_XLSX)


def main() -> None:
    rows = build_review()
    write_rows(OUTPUT_CSV, rows)
    write_xlsx(rows)
    summary = {
        "input": str(INPUT_CSV),
        "output_csv": str(OUTPUT_CSV),
        "output_xlsx": str(OUTPUT_XLSX),
        "total_rows": len(rows),
        "review_conclusion_counts": Counter(row["复核结论"] for row in rows),
        "system_ready_count": sum(1 for row in rows if row["是否建议接入系统"] == "是"),
        "mapped_scene_counts": Counter(row["映射训练场景"] for row in rows),
    }
    SUMMARY_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
